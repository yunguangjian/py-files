# -*- coding: utf-8 -*-
import pyautogui
import time
from paddleocr import PaddleOCR
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
import win32gui
import pyperclip
import win32con
import sys
try:
    # 通常为激活ERP系统句柄至最前来实现切换窗口，如果系统没打开，则运用alt+tab来切换窗口
    pycharm = win32gui.GetForegroundWindow()#换取pycharm句柄
    win32gui.ShowWindow(pycharm, win32con.SW_SHOWMINNOACTIVE)#pycharm最小化
    jb=win32gui.FindWindow(0,'基层店管理系统')       #引号内输入要切换的窗口名称
    win32gui.ShowWindow(jb,1)                      #应对窗口最小化情况
    win32gui.SetForegroundWindow(jb)                #激活窗口至最前1，仅能激活
    win32gui.SetWindowPos(jb,win32con.HWND_TOP,448,137,1016,775,win32con.SWP_SHOWWINDOW)        #调整窗口在屏幕位置和大小
except:
    pyautogui.alert('ERP系统没开，终止程序')
    sys.exit()
pyautogui.alert('请检查首列是否是“书名”，检查完毕后按“确定”，进行下一步操作')
pyautogui.alert('第一次运行，请确认截图范围！请把鼠标放到“截图右下角”位置，并按“空格”确认')
x,y=pyautogui.position()
pyautogui.alert('第一次运行，请确认上翻页点击位置！请把鼠标放到“点击”位置，并按“空格”确认')
p,q=pyautogui.position()
ocr = PaddleOCR(lang="ch", show_log=False)
wb=openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\学校发货册数条目统计.xlsx')
sht=wb['Sheet1']
cc=sht.max_row
for cx in range(2,cc+1):
    pyautogui.click(x=1205, y=341)
    for _ in range(4):
        pyautogui.typewrite(['tab'])
    pyperclip.copy(sht.cell(cx,1).value)
    pyautogui.hotkey('ctrl', 'v')
    for _ in range(4):
        pyautogui.typewrite(['tab'])
    pyautogui.typewrite('01')
    for _ in range(2):
        pyautogui.typewrite(['tab'])
    time.sleep(1)
    pyautogui.typewrite(['enter'])
    # pyautogui.alert('123')
    time.sleep(7)#根据查询完后出现数据的间隔时间进行调整
    pyautogui.moveTo(1190, 370)  # “原免费册数”向前移，相邻到“书名”
    pyautogui.mouseDown()
    pyautogui.moveTo(987, 378)
    pyautogui.mouseUp()
    time.sleep(1)
    pyautogui.mouseDown()  # 扩大“书名”列宽
    pyautogui.moveTo(1100, 378)
    pyautogui.mouseUp()

    time.sleep(1)
    screenshot = pyautogui.screenshot(region=(670, 385, x - 670, y - 385))
    screenshot.save('screenshot.png')#非必要，不过可以截图看看截图对不对
    jietu = ocr.ocr(np.array(screenshot))

    if sht.cell(cx,2).value not in wb.sheetnames:
        wb.create_sheet(sht.cell(cx,2).value)

    wb[sht.cell(cx,2).value].cell(1,1,'书名')
    wb[sht.cell(cx, 2).value].cell(1, 2, '原免费册数')
    wb[sht.cell(cx, 2).value].cell(1, 3, '前二合一')
    wb[sht.cell(cx, 2).value].cell(1, 4, '排序')

    yushu=sht.cell(cx,3).value%23
    shang=int(sht.cell(cx,3).value/23)
    lb=[]
    for a in jietu:
        for b in a:
            # c=b[1][0].replace(" ","")
            c = b[1][0].replace("（", "(").replace("）", ")")
            lb.append(c)
    if len(lb)%2==1:
        lb.pop()
    js=2
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for a in range(0,len(lb),2):
        wb[sht.cell(cx, 2).value].cell(js, 1, lb[a])
        try:
            wb[sht.cell(cx, 2).value].cell(js, 2, int(lb[a + 1]))
        except:
            wb[sht.cell(cx, 2).value].cell(js, 2, lb[a + 1])
            wb[sht.cell(cx, 2).value].cell(js, 2).fill=fill
            sht.cell(cx, 5, '识别有问题')
        wb[sht.cell(cx, 2).value].cell(js, 3, lb[a] + lb[a + 1])  # 新加，注意运行是否有问题
        wb[sht.cell(cx, 2).value].cell(js, 4, js-1)
        js=js+1
    if yushu==0:
        for _ in range(shang-1):
            pyautogui.typewrite(['pagedown'])
            screenshot = pyautogui.screenshot(region=(670, 385, x - 670, y - 385))
            time.sleep(0.5)
            screenshot.save('screenshot.png')  # 非必要，不过可以截图看看截图对不对
            jietu = ocr.ocr(np.array(screenshot))
            lb = []
            for a in jietu:
                for b in a:
                    c = b[1][0].replace("（", "(").replace("）", ")")
                    lb.append(c)
            if len(lb) % 2 == 1:
                lb.pop()
            for a in range(0, len(lb), 2):
                wb[sht.cell(cx, 2).value].cell(js, 1, lb[a])
                try:
                    wb[sht.cell(cx, 2).value].cell(js, 2, int(lb[a + 1]))
                except:
                    wb[sht.cell(cx, 2).value].cell(js, 2, lb[a + 1])
                    wb[sht.cell(cx, 2).value].cell(js, 2).fill = fill
                    sht.cell(cx, 5, '识别有问题')
                wb[sht.cell(cx, 2).value].cell(js, 3, lb[a]+lb[a + 1])#新加，注意运行是否有问题
                wb[sht.cell(cx, 2).value].cell(js, 4, js - 1)
                js = js + 1
    else:
        for fy in range(shang):
            pyautogui.typewrite(['pagedown'])
            if fy+1==shang:
                pyautogui.moveTo(p,q)
                pyautogui.click()
                pyautogui.click()
            screenshot = pyautogui.screenshot(region=(670, 385, x - 670, y - 385))
            time.sleep(0.5)
            screenshot.save('screenshot.png')  # 非必要，不过可以截图看看截图对不对
            jietu = ocr.ocr(np.array(screenshot))
            lb = []
            for a in jietu:
                for b in a:
                    c = b[1][0].replace("（", "(").replace("）", ")")
                    lb.append(c)
            if len(lb) % 2 == 1:
                lb.pop()
            if fy+1==shang:
                lb=lb[(23-yushu)*2:]
            for a in range(0, len(lb), 2):
                wb[sht.cell(cx, 2).value].cell(js, 1, lb[a])
                try:
                    wb[sht.cell(cx, 2).value].cell(js, 2, int(lb[a + 1]))
                except:
                    wb[sht.cell(cx, 2).value].cell(js, 2, lb[a + 1])
                    wb[sht.cell(cx, 2).value].cell(js, 2).fill = fill
                    sht.cell(cx, 5, '识别有问题')
                wb[sht.cell(cx, 2).value].cell(js, 3, lb[a] + lb[a + 1])  # 新加，注意运行是否有问题
                wb[sht.cell(cx, 2).value].cell(js, 4, js - 1)
                js = js + 1
    wb.save(r'C:\Users\Administrator\Desktop\学校发货册数条目统计.xlsx')